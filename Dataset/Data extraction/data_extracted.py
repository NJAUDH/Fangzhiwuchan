import json
import re
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
from tqdm import tqdm


def extract_nested_entities(text):
    pattern = r'\[(.*?)\]'
    stack = []
    entities = []
    clean_text = ""

    for pos, char in enumerate(text):
        if char == '[':
            stack.append((pos, ""))
        elif char == ']':
            if stack:
                start, entity_name = stack.pop()
                entity_full_text = text[start + 1:pos]

                if '|' in entity_full_text:
                    entity_name, entity_type = entity_full_text.rsplit('|', 1)
                else:
                    entity_name, entity_type = entity_full_text, None

                entities.append({
                    "entity_name": entity_name.strip(),
                    "entity_type": entity_type.strip() if entity_type else "",
                    "start_pos": start - len(clean_text),
                    "end_pos": start - len(clean_text) + len(entity_full_text.strip()) - 1
                })
        else:
            if stack:
                stack[-1] = (stack[-1][0], stack[-1][1] + char)
            else:
                clean_text += char


    annotated_description = re.sub(r'[^\u4e00-\u9fff]', '', text)
    return entities, clean_text, annotated_description


def process_xlsx_to_json(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active
    results = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        annotated_description = row[6]
        if annotated_description:
            id = row[0]
            entities, original_description, clean_annotated_description = extract_nested_entities(annotated_description)
            results.append({
                "id": id,
                "original_description": annotated_description,
                "annotated_description": clean_annotated_description,
                "extracted_entities": entities
            })

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=4, ensure_ascii=False)

    return output_file


def clean_entities(data):
    for item in data:
        for entity in item['extracted_entities']:
            entity['entity_name'] = re.sub(r'[^\u4e00-\u9fff]', '', entity['entity_name'])
    return data


def process_json(input_json, output_json):
    with open(input_json, 'r', encoding='utf-8') as file:
        data = json.load(file)

    cleaned_data = clean_entities(data)

    with open(output_json, 'w', encoding='utf-8') as file:
        json.dump(cleaned_data, file, ensure_ascii=False, indent=4)

    return output_json


def clean_chinese_text(text):
    return re.sub(r'[^\u4e00-\u9fff]', '', text)


def find_entity_positions(clean_description, entities):
    for entity in entities:
        clean_entity_name = clean_chinese_text(entity['entity_name'])
        matches = [(m.start(), m.end()-1) for m in re.finditer(re.escape(clean_entity_name), clean_description)]
        entity['start_pos'], entity['end_pos'] = matches[0] if matches else (-1, -1)


def process_single_item(item):
    clean_description = clean_chinese_text(item['annotated_description'])
    find_entity_positions(clean_description, item['extracted_entities'])
    return item


def process_json_entities(input_json, output_json):
    with open(input_json, 'r', encoding='utf-8') as file:
        data = json.load(file)

    with ThreadPoolExecutor() as executor:
        data = list(tqdm(executor.map(process_single_item, data), total=len(data), desc="Processing items"))

    with open(output_json, 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)

    return output_json


input_xlsx_path = 'original_data.xlsx'
output_json_path1 = 'biaozhu1.json'
output_json_path2 = 'biaozhu2.json'
output_json_path3 = 'biaozhu3.json'


process_xlsx_to_json(input_xlsx_path, output_json_path1)


process_json(output_json_path1, output_json_path2)


process_json_entities(output_json_path2, output_json_path3)

print(f"最终处理完成，JSON文件生成: {output_json_path3}")
